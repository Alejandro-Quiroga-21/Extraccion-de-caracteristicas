import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from scipy import signal
import pywt
from scipy.stats import entropy
from collections import deque 
from collections import defaultdict
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.model_selection import GridSearchCV, StratifiedKFold, cross_val_score, RepeatedStratifiedKFold
from sklearn.svm import SVC
from openpyxl import load_workbook, Workbook
from sklearn.metrics import accuracy_score, recall_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

pipeline = Pipeline([
    ("scaler", StandardScaler()),
    ("svm", SVC(kernel="linear", class_weight="balanced", random_state=42))
])

dictionary = np.load('dictionary_512.npy', allow_pickle=True).item()
print(type(dictionary))
volunteers = sorted(dictionary.keys())
print('Volunteers:',volunteers)

df_info = pd.read_excel('C:/Users/Usuario/Documents/GitHub/Data_base_processing/Database_Info.xlsx')
upper_dominant = (df_info["Upper Dominant Laterality"]).tolist()
lower_dominant = (df_info["Lower Dominant Laterality"]).tolist()
print('Upper dominant=',upper_dominant)
print('Lower dominant=',lower_dominant)

feet_channels = ['CZ', 'C4', 'C3', 'FZ', 'PZ']

trials, channels, samples = dictionary['S01']['Lower']['Task1']['X']['Rest'].shape
print(f"Trials={trials}|Channels={channels}|Samples={samples}")

### Laplacian filter
def laplacian_filter(senal):
  samples, channels, trials = senal.shape
  laplacian = np.empty((samples, trials), np.float64) 

  for i in range(trials):
      for k in range(samples):
          sample_sum = 0
          for j in range(1, channels):
              sample_sum += senal[k, j, i] 
          laplacian[k, i] = senal[k, 0, i] - (sample_sum / (channels - 1))  
  return laplacian

def PSD_rithms(signals, fs=512, nperseg=512, overlap=256, lfreq=8, hfreq=30):
    trials, samples = signals.shape
    psd_values = []
    for trial in range(trials):
        freqs, psd = signal.welch(signals[trial], fs, nperseg=nperseg, noverlap=overlap, window='hamming', scaling='density')
        # Filtrar valores de PSD en el rango de 8 a 30 Hz
        indices = np.where((freqs >= lfreq) & (freqs <= hfreq))
        filtered_psd = psd[indices]  # Valores PSD en el rango deseado
        
        psd_values.append(filtered_psd)
    
    psd_values = np.array(psd_values)
    filtered_freqs = freqs[indices]
    return psd_values, filtered_freqs

def cross_valid_lda(X,y):
    clf = LinearDiscriminantAnalysis(shrinkage='auto', solver='lsqr')

    # RepeatedStratifiedKFold: 4 folds, repetido 10 veces
    cv = RepeatedStratifiedKFold(n_splits=4, n_repeats=10, random_state=42)

    acc_scores = cross_val_score(clf, X, y, cv=cv, scoring='accuracy')
    tpr_scores = cross_val_score(clf, X, y, cv=cv, scoring='recall')  # recall = sensibilidad

    return (np.mean(acc_scores), np.std(acc_scores), np.mean(tpr_scores), np.std(tpr_scores) ,clf)

def cross_valid_svm(X,y):
     # Definir modelo base
    svm = SVC(kernel='rbf', random_state=42)
    param_grid = {
    'C': [0.1, 1, 10, 100],  # Diferentes valores de C
    'gamma': [0.001, 0.01, 0.1, 1, 'scale', 'auto']  # Diferentes valores de gamma
    }

    # svm = SVC(kernel='linear', random_state=42)
    # # Búsqueda de hiperparámetro C
    # param_grid = {'C': [0.1, 1, 10, 100, 1000]}

    grid_search = GridSearchCV(estimator=svm,
                               param_grid=param_grid,
                               cv=4,            # CV interna para elegir C
                               scoring='accuracy',
                               n_jobs=-1)
    grid_search.fit(X, y)

    print("Mejores parámetros:", grid_search.best_params_)
    print("Mejor score (accuracy) durante la búsqueda:", grid_search.best_score_)

    # Mejor modelo encontrado
    clf = grid_search.best_estimator_

    # Validación cruzada repetida (5 folds × 10 repeticiones = 50 scores)
    cv = RepeatedStratifiedKFold(n_splits=4, n_repeats=10, random_state=42)

    acc_scores = cross_val_score(clf, X, y, cv=cv, scoring='accuracy')
    tpr_scores = cross_val_score(clf, X, y, cv=cv, scoring='recall')

    # Devolvemos promedio y desviación estándar (muy útil para comparar métodos)
    return (np.mean(acc_scores), np.std(acc_scores),
            np.mean(tpr_scores), np.std(tpr_scores),
            clf)

method_used = {'acc_lda':[], 'acc_std_lda':[] ,'tpr_lda':[], 'tpr_std_lda':[],
               'acc_svm':[], 'acc_std_svm':[], 'tpr_svm':[], 'tpr_std_svm':[],
          }

sheets = ['T1-LDA','T1-SVM']

### main

for volunteer in range(1,31):

    if volunteer < 10:
        ini = 'S0'
    else:
        ini = 'S'

    X = dictionary[ini+str(volunteer)]['Lower']['Task1']['X']
    Y = dictionary[ini+str(volunteer)]['Lower']['Task1']['Y']

    ### Laplacian

    laplace_left = laplacian_filter(X['Left'])
    trials, samples = laplace_left.shape
    #print(f"Left Trials={trials}|Samples={samples}")
    laplace_right = laplacian_filter(X['Right'])
    trials, samples = laplace_right.shape
    #print(f"Right Trials={trials}|Samples={samples}")
    laplace_rest = laplacian_filter(X['Rest'])
    trials, samples = laplace_rest.shape
    #print(f"Rest Trials={trials}|Samples={samples}")

    ### PDS Welch
    PSD_left, freq = PSD_rithms(laplace_left)
    PSD_right, freq = PSD_rithms(laplace_right)
    PSD_rest, freq = PSD_rithms(laplace_rest)

    ### Data classifier
    if lower_dominant[volunteer-1] == 'Right':
        index = np.random.choice(PSD_rest.shape[0], 40, replace=False) 
        selection = PSD_rest[index]
        X = np.concatenate((PSD_right, selection), axis=0)
    else:
        index = np.random.choice(PSD_rest.shape[0], 40, replace=False) 
        selection = PSD_rest[index]
        X = np.concatenate((PSD_left, selection), axis=0)

    y = np.concatenate((np.ones(40),np.zeros(40)), axis=0)

    print('Tamaño X:', X.shape, 'Tamaño y:', y.shape)

    ### Classifiers
    ### LDA
    acc_m, acc_std, tpr_m, tpr_std, clf = cross_valid_lda(X,y)
    method_used['acc_lda'].append(acc_m)
    method_used['acc_std_lda'].append(acc_std)
    method_used['tpr_lda'].append(tpr_m)
    method_used['tpr_std_lda'].append(tpr_std)

    ## SVM
    acc_m, acc_std, tpr_m, tpr_std, clf = cross_valid_svm(X,y)
    method_used['acc_svm'].append(acc_m)
    method_used['acc_std_svm'].append(acc_std)
    method_used['tpr_svm'].append(tpr_m)
    method_used['tpr_std_svm'].append(tpr_std)

### Save 

def save_results(sheets, data):
    excel_filename = 'Results_feet_512_T1.xlsx'
    try:
        wb = load_workbook(excel_filename)
    except FileNotFoundError:
        wb = Workbook('Results_feet_512_T1.xlsx')
        ws = wb.active  # Obtener la hoja activa por defecto

    data_keys = list(data.keys())
    print(data_keys)


    for i in range(0,len(data_keys),4):
        results = pd.DataFrame({
            'Accuracy': data[data_keys[i]],
            'Acc_std': data[data_keys[i+1]],
            'Recall': data[data_keys[i+2]],
            'Recall_std': data[data_keys[i+3]],
            })
        
        # Acceder a la hoja 
        if sheets[int(i/4)] not in wb.sheetnames:
            ws = wb.create_sheet(sheets[int(i/4)])
        else:
            ws = wb[sheets[int(i/4)]]

        # Especificar la celda de inicio (C6)
        start_row = 6
        start_col = 3  # Columna C

        # Escribir los encabezados en la fila 5
        ws.cell(row=5, column=start_col, value='Accuracy')
        ws.cell(row=5, column=start_col + 1, value='std')
        ws.cell(row=5, column=start_col + 2, value='TPR (Recall)')
        ws.cell(row=5, column=start_col + 3, value='std')

        # Escribir los datos del DataFrame a partir de la celda C6
        for i, row in enumerate(results.itertuples(), start=0):
            ws.cell(row=start_row + i, column=start_col, value=row.Accuracy)
            ws.cell(row=start_row + i, column=start_col + 1, value=row.Acc_std)  
            ws.cell(row=start_row + i, column=start_col + 2, value=row.Recall)
            ws.cell(row=start_row + i, column=start_col + 3, value=row.Recall_std)

        # Guardar el archivo Excel
        wb.save(excel_filename)

save_results(sheets, method_used)
